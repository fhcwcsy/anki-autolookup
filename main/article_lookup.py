"""
This program will create a lookup window where the user can enter/paste an article 
and choose the difficulty of words he wants to be looked up. The program will 
find the words in the article that match the requirements and list in the wordlist
window on his right. Then, user can choose the words he want to add to anki.

The _difficulty_ of a word is defined as follows: We use a wordlist from  
[NGSL](http://www.newgeneralservicelist.org/), which contains the words used
in fictions, journals, TV subtitles, etc., and their times of being used. Hence,
the difficulties of a word is defined by the frequency of being used. If a word
is more oftenly used, it is considered to be less difficult, and vice versa. If
a word is not in the list, it is considered to be too difficult ore rarely used
for a foreign English learner to learn. The frequency used in this program is
normalized by dividing the actual count of usage by the number of times used of
the most frequently used word (so the frequency is between 0 and 1). This 
program reads the text paste in the textbox, determine the difficulty of each 
word by looking up in the local wordlist, then lookup each word in the Cambridge
dictionary. 

The required modules are:

- tkinter
- re (regular expression)
- openpyxl 
- wordlist_cls: This module is created by us.
"""

import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
import re
import wordlist_cls


class ArticleRecognitionWindow:
    """
    The class controlling the article lookup feature.

    Attributes:
        difficulty: The (normalized) maximum difficulty of the word to be looked
            up. The lower this value is, the less word (keeping only the most
            difficult ones) will be looked up.
    """

    def __init__(self):
        """
        Create two windows. 

        One is `_inputWindow` where the users enter the article and determine
        the difficulty of words they want to be looked up.
    
        The other is `_wordwindow` where we show the the difficult words we
        find in the article. Users can select which words they want to make
        vocabulary cards here.
        """
        self._inputWindow = tk.Toplevel()
        self._wordwindow = tk.Toplevel()
        self._inputWindow.title('Article Lookup')
        self._inputWindow.geometry('800x600+20+20')
        self.difficulty = 4e-5

        # self._inputWindow.configure(background='black')
        # text_lookup_frame = tk.Frame(self._inputWindow)
        # text_lookup_frame.pack(side=tk.TOP)
        text_lookup_label = tk.Label(self._inputWindow,
                text='Article to lookup: ', bg='white', font=('Arial',
                18))
        text_lookup_label.pack(side=tk.TOP)
        difficultyFrame = tk.Frame(self._inputWindow)
        difficultyPrompt = tk.Label(difficultyFrame, text='Difficulty:')
        self._difficultyStr = tk.StringVar(value='4e-5') 
        difficultyBox = tk.Entry(difficultyFrame, width = 25,
                textvariable=self._difficultyStr)
        difficultyFrame.pack()
        difficultyPrompt.grid(column=1, row=1, pady=(15, 0))
        difficultyBox.grid(column=2 ,row=1, pady=(15, 0), sticky='w')
        difficultyHint = tk.Label(difficultyFrame, text=
            '1=lookup every word, 1e-6=lookup only the most difficult words.')
        difficultyHint.grid(column=1, row=2, columnspan=2, pady=(0, 15))
        self._lookupButton = tk.Button(self._inputWindow, text='Lookup!'
                , command=self.lookup)
        self._lookupButton.pack()
        self.outputbox = tk.Text(self._inputWindow, width=110,
                                 height=45)
        self.outputbox.pack()
        self._wordwindow.geometry('300x600+820+20')

        self._wordwindow.title('Words to be added')
        self.wlist = wordlist_cls.WordlistWindow(self._wordwindow,
                self.quitWindow, bg='#444444')
        self.wlist.pack(expand='true', fill='both')

        self._inputWindow.mainloop()

    def quitWindow(self):
        """
        Destroy/quit the windows after using it.

        Args:
            None
            
        Return:
            None
            
        Raise:
            None
        """
        self._inputWindow.destroy()
        self._inputWindow.update()
        self._inputWindow.quit()
        self._wordwindow.destroy()
        self._wordwindow.destroy()
        self._wordwindow.quit()

    def lookup(self):
        """
        This function will first get the difficulty users set. Use it to find
        the difficult words in the article, and add it to the `_wordwindow`.

        Args:
            None
    
        Return:
            None
    
        Raise:
            raise Exception('InvalidDifficulty') if self.difficulty is not 
                between 0 and 1.
        """
        try:
            self.difficulty = float(self._difficultyStr.get())
            # print(self.difficulty)
            if not ( 0 <= self.difficulty <= 1):
                raise Exception('InvalidDifficulty')
        except Exception as e:
            messagebox.showerror('Error', 'Invalid difficulty!')
            return
            # raise e
        text = self.outputbox.get('1.0', 'end')
        worddic = self._getWordlist()
        wordset = self._getArticle(text)
        # print(wordset)
        wordlist = self._setLookup(wordset, worddic)
        for word in wordlist:
            self.wlist.newWord(word)

    def _getWordlist(self):
        """
        Read the excel sheet (the local dictionary with frequencies of the words)
        and convert it into a dictinary with the key being the words and and 
        the value being its frequency/max_frequency.

        Args:
                None

        Return: 
                A dictionary

        Raise:
                None
        """
        WBPATH = './dic.xlsx'
        ws = load_workbook(WBPATH)['Freq']
        freq_max = ws['B2'].value
        wordlist = {}
        for i in range(2, 22233):
            wordlist[ws['A{}'.format(i)].value] = (
                    ws['B{}'.format(i)].value / freq_max )
        return wordlist

    def _getArticle(self, text):
        """
        Read the text, cut it into words, then add them into a set then return.

        Args:
            text: The article we want to anaylze.
            
        Return: 
            A set containing the words.
    
        Raise:
            None
        """
        words = set()
        splitchars = '[\s+,."]'
        words.update(re.split(splitchars, text.lower()))
        return words

    def _dictLookup(self, word, d):
        """
        Local wordlist lookup. Check if the word is in the wordlist and get 
        its frequency. Also check the stripped words if the word matches any
        common suffixes.
    
        Args:
            word: the word to look up
            d: The dictionary to look up
    
        Return: 
            If any result is found, then return a tuple of (word, frequency).
            If the word (or any of the stripped version) is not in the
            list, then return None.
            
        Raise:
            None
        """
        # print(d)
        w = word
        result = d.get(w, 0)
        if result:
            # print(result)
            return (word, result)

        if w[-1] == 's':
            w = w.strip('s')
            result = d.get(w, 0)
            if result:
                return (w, result)
            w = word

        if w[-2:] == 'es':
            w = w.strip('es')
            result = d.get(w, 0)
            if result:
                return (w, result)
            w = word

        if w[-2:] == 'ed':
            w = w.strip('ed')
            result = d.get(w, 0)
            if result == 0:
                w = w.strip('d')
                result = d.get(w, 0)
            if result:
                return (w, result)
            w = word
        return None

    def _setLookup(self, s, dictionary):
        """Look up each word in the set in the local wordlist.
        
        If the word is found and its frequency is below the self.difficulty,
        then recognize it as difficult word. Stopwords are removed.

        Args:
            s: a set containing the words to be looked up.
            dictionary: the dictionary with the key be the words and the value
                be its frequency.
        return:
            a list containing the difficult words found in the local wordlist.
        
        Raise:
            None
        """

        stopWords = [
            '',
            'i\'m'
            'is',
            'are',
            'were',
            'was',
            'have',
            'has',
            'a',
            's',
            ]
        vocabList = []
        for be in stopWords:
            if be in s:
                s.remove(be)
        # print(s)
        for word in s:
            lookupResult = self._dictLookup(word, dictionary)
            if lookupResult == None:
                continue
            # print(lookupResult[1], self.difficulty)
            if 0 < lookupResult[1] < self.difficulty:
                vocabList.append(word)
        return vocabList


if __name__ == '__main__':
    done = ArticleRecognitionWindow() 
    # worddic = self._getWordlist()
