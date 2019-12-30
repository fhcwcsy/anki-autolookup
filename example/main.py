from openpyxl import load_workbook
import re
import requests
import urllib.request 
from bs4 import BeautifulSoup 
import argparse
from argparse import RawTextHelpFormatter


# Parser settings
parser = argparse.ArgumentParser(prog='AutoLookup', usage='$ python main.py [optional arguments]', description=
        'Goal: Look up difficult words from a text file in an online dictionary. A local wordlist from \n' + 
        'http://www.newgeneralservicelist.org/ is used, which contains the top words used in American TV\n' + 
        'subtitles and other sources, and their frequency of being used. This wordlist is used to define \n' + 
        'the difficulty of a word. The frequency used in this program is normalized by dividing the real \n' +
        'frequency by the max frequency. This program reads a text file containing some English words, \n' +
        'determine its difficulty by looking up in the local wordlist. If the (normalized) frequency is \n' +
        'below a certain value, which can be specified by users and has a default value of 4e-5, then the \n' +
        'program will look it up in the specified online dicitonary. Currently only lexico.com and \n' +
        'dictionary.com is provided, but the program can be easily modified to add other sources. The \n' +
        'source has better lookup results as its webpage is a lot more simple for scrapers, so is set \n' +
        'to be the default value of the source. The scraper only find the first meaning found in the \n' + 
        'dictionary, but, again, it can be easily modified to find more. For more details about arguments, \n' +
        'see the Arguments section. For details about the functions and how they works, see the docstrings \n' +
        'in the code.\n\n' + 
        'The modules used by this program are:\n' + 
        '- openpyxl\n' + 
        '- regular expression\n' + 
        '- requests\n' + 
        '- urllib\n' + 
        '- BeautifulSoup\n' + 
        '- argparse\n\n' + 
        'You may need to pip3/conda/apt-get/other_package_managers install something if you are missing \n' + 
        'any one of these.'
        , formatter_class=RawTextHelpFormatter ) 
parser.add_argument('--path', '-p', default='./text.txt', type=str, dest='path', help='Path to the text file. The default value is ./text.txt') 
parser.add_argument('--source', '-s', dest='source', type=str, default='lexico.com', help='Indicate the dictionary source. Currently only dictionary.com and lexico.com \nare provided. The latter has much better results. The default value is lexico.com')
parser.add_argument('--threshold', '-t', dest='threshold', type=float, default=4e-5, help='A float between 0 and 1. Only more difficult words will be looked up if a \nlower value is specified. The default value is 4e-5.')
args = parser.parse_args()

# Constants
WBPATH = './dic.xlsx'
THRESHOLD = args.threshold
if THRESHOLD > 1 or THRESHOLD < 0:
    raise(Exception('Invalid threshold specified'))
stopWords = ['', 'is', 'are', 'were', 'was', 'have', 'has', 's']
splitchars = '[\s+,.\'"]'
SOURCE = args.source

#functions

def getWordlist():
    """Read the excel sheet and convert it into a dictinary containing the words and its frequency/max_frequency.
    No arguments
    return: a dictionary
    """

    ws = load_workbook(WBPATH)['Freq']
    freq_max = ws['B2'].value
    wordlist = {}
    for i in range(2,22233):
        wordlist[ws['A{}'.format(i)].value] = ws['B{}'.format(i)].value / freq_max
    return wordlist

def getArticle(path):
    """Read the text file, cut it into words, then add them into a set.
    Arguments:
    path: A string specifying the path to the text file
    return: a set
    """
    f = open(path)
    words = set()
    for line in f:
        words.update(re.split(splitchars, line.strip().lower()))
    return words

def onlineLookup(word):
    """Look up a word in the specified dictionary source. The word provided must be found in the local wordlist. If the lookup fails, then return a tuple of (word, None).
    Arguments: 
    word: the word to look up
    return: A tuple containing the word and its definition found
    """
    if SOURCE == 'lexico.com':
        try:
            url = 'https://www.lexico.com/en/definition/' + word 
            response = requests.get(url)
            soup = BeautifulSoup(response.text, "html.parser") 
            results = soup.find('span', 'ind').text
        except Exception as e:
            return word, None
        
    elif SOURCE == 'dictionary.com':
        try:
            url = 'https://www.dictionary.com/browse/' + word 
            response = requests.get(url)
            soup = BeautifulSoup(response.text, "html.parser") 
            results = soup.find('span', 'e1q3nk1v4').text
        except Exception as e:
            return word, None


    else:
        raise(Exception('SourceError: Unknown dictionary provided.'))
    return word, results
    
def dictLookup(word, d):
    """Local wordlist lookup. Check if the word is in the wordlist and get its frequency. Also check the stripped words if the word matches any common suffixes.
    Arguments:
    word: the word to look up
    d: The dictionary to look up
    return: If any result is found, then return a tuple of (word, frequency). If the word (or any of the stripped version) is not in the list, then return None.
    """
    w = word
    result = d.get(w, 0)
    if result:
        return word, result

    if w[-1] == 's':
        w = w.strip('s')
        result = d.get(w, 0)
        if result:
            return w, result
        w = word

    if w[-2:] == 'es':
        w = w.strip('es')
        result = d.get(w, 0)
        if result:
            return w, result
        w = word

    if w[-2:] == 'ed':
        w = w.strip('ed')
        result = d.get(w, 0)
        if result == 0:
            w = w.strip('d')
            result = d.get(w, 0)
        if result:
            return w, result
        w = word

    return None

def setLookup(s, dictionary):
    """Look up each word in the set in the local wordlist. If the word is found and its frequency is below the THRESHOLD, then do online lookup. Stopwords are removed.
    Arguments:
    s: a set containing the words to be looked up.
    dictionary: the dictionary with the key be the words and the value be its frequency.
    return: a list containing tuples of (difficult_words, definition)
    """
    vocabList = []
    for be in stopWords:
        if be in s:
            s.remove(be)
    for word in s:
        lookupResult = dictLookup(word, dictionary)
        if lookupResult == None:
            continue
        if 0 < lookupResult[1] < THRESHOLD:
            vocabList.append(onlineLookup(lookupResult[0]))

    return vocabList

if __name__ == "__main__":
    path = args.path
    d = getWordlist()
    setOfWords = getArticle(path)
    result = setLookup(setOfWords, d)
    for word, definition in result:
        print(word,': ', definition)
